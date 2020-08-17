
#####
pst= input("what is the positiona being applied for? ")
cmp= input("what is the comapny targeted name? ")
cvr_ltr= f' Hi {cmp}, I am applying for the {pst} position.'
print(cvr_ltr)
######

name= "Mohamed Mahmoud Bayoumi"
print(name.find("m"))


print(name.count("m"))
print(10**3)
import math

math.ceil(2.9)
]
math.frexp(1**15)
math.fsum([0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1,0.1])
math.sqrt(26)
float in (type(-2.354))
is_integer
isinstance(2.34,float)
2.34.is_integer()



######

Temp = 25
if Temp > 30:
    print("It's a hot day")
elif Temp < 10:
    print("It's a cold day")
else:
    print("It's a wonderful day")


##########
password = "2195&Omar"

print(("!" or "&") in password)
######
Weight= input("what's your weight? ")
unit= input("What's the unit Kg (K) or Ib (I)?")
if unit.upper() == "K":
    weight_Ib = int(Weight) *2.20462
    print(f"weight in Kg =  {weight_Ib}")
elif unit.upper() == "I" :
    weight_Kg = int(Weight) * 0.453592
    print(f"weight in Kg =  {weight_Kg}")
else:
    print(" error please reenter the units abbreviation and try again")
#####
i=1
while i <= 3:
    G= int(input( "Guess! "))
    if G == 9:
        print("WOW! You Win!")
        break
    i=i+1
else:
    print("Try Again")
######
command = ""
started =  False
while command != "quit":
    command = input("> ").lower()
    if command == "start":
        if started:
            print("car already running")
        else:
            started = True
            print("car started...")

    elif command == "stop":
        print("car stopped")

items = [1,2,3,4,5,6]
total = 0
total2 = [0,0,0,0,0,0]
for i in items:
    total2[i-1]=items[i-1]
    total+= i
print(f' total= {total}')

print(total2)
items [0]
total2[2]=items[2]

######
letter_L = [2,2,2,2,5]
for Z in letter_L:
    output = ""
    for X in range(Z):
        output+= "X"
    print(output)
#####
f_large=[45,324,45,1054,654]
largest = f_large [0]
for X in f_large:
    if X > largest:
        largest = X
print(f'largest= {largest}')


#####
matrix = [[1,2,3]
    ,[4,5,6]
    ,[7,8,9]]

print(matrix)


for z in matrix:
    for x in z:
        print(x)

#### remove duplicate
numb = [1,5,5,8,63,54,4,6,2,4,3,3]
numb.sort()
unique = []
for X in numb:
    if X not in unique:
        unique.append(X)
print(unique)



p_numb = int(input("phone "))
number_alpha = { 1:"ONE ", 2:"two ", 3:"three ", 4:"four ",5:"five ",6:"six ",7:"seven ",8:"eight ",9:"nine ",0:"zero "}
phone_number_text = ""
for X in p_numb:
    phone_number_text+= number_alpha[(X)]
print(phone_number_text)


p_numb = int("1,2")
p_numb= input("phone? ")
print(p_numb)
print(type(p_numb))
for c in p_numb:
    print(c)

number_alpha.get(1,"k")


#######
def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print("welcome abroad")


print("start")
greet_user(input("what's your first name? "),input("what's your last name? "))
print("finish")

def square(number):
    print(number*number)
    return


print(square(3))
###### To anticipate error (try:   except)
try:
    age= int(input("Age> "))
    print(f'birth year is {2020-age}')
except ValueError:
    print("Age MUST be a numerical format")
###
class Person:
    def __init__(self, name, language):
        self.n = name
        self.l= language
    def introduce(self):
        print(f'Hi, I am {self.n}, and I speak {self.l}')


id_1= Person(input("what is your name? "), input("what is your native language? "))
id_1.introduce()


######3
import converter

###OR
from converter import
print(converter.Kg_to_Ib(80))
##3
from converter import find_max
find_max([5415,465,4,54,54,65,465,4])


####
import ecommerce.shipping
ecommerce.shipping.calc_ship()
from ecommerce.shipping import calc_ship

#####
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]
#OR
cell= sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value *0.9
    corrected_price_cell = sheet.cell(row,4) # to create a new cell
    corrected_price_cell.value = corrected_price
corrected_price_cell= sheet.cell (1,4)
corrected_price_cell.value = "updated price"

values =  Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col= 4, max_col = 4)# to select the cells of interest (mouse drag)
chart = BarChart()

chart.add_data(values)
sheet.add_chart(chart,"e2")
wb.save("transactions2.xlsx")
